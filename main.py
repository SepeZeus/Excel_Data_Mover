#! /usr/bin/env python3
from package.misctools import MiscTools as mt
import openpyxl
from openpyxl import load_workbook
import itertools
import string
import shutil

filePath = input("Drag file here: ")
filePath = filePath.strip('\"') #gets rid of quotation marks from start and end of path name

wb = load_workbook(filePath)
ws = wb.active

def excel_data_mover(): #retrieves all data from specific excel columns
    colData = []
    colTake= input("Column(s) to get data from: ").split()
    dup = input("Column(s) data to duplicate: ").split()
    if (len(colTake) > 1):
        zipCols = input("Which columns to zip: ").split()

    def col_extractor():
        data = []
        for colNum in colTake:
            colChar = colNum
            colNum = mt.col_to_num(colNum)
            #gets column data from excel file
            for rowCells in ws.iter_cols(min_col=colNum, max_col=colNum, min_row=2): 
                for cell in rowCells:
                    if(cell.value):
                        data.append(cell.value)

                    elif(cell.value == None):
                        data.append('-')

            #checks for data to duplicate 
            if(colChar in dup):
                data = mt.list_value_duplicator(data, 2)    
                
            colData.append(data) 
            data = []


    def zip_cols(zipCols): #zips data and adds it back to previous data
    
        zippedData = []
        disposables = []
        colChars = colTake
        skipChar = False
        skipInd = False
        nPrev = 0
    
        def list_zipper(l1, l2): #creates one long list out of two separate lists
            zipped = itertools.zip_longest(l1, l2)
            zipped = list(itertools.chain.from_iterable(zipped))
            return zipped
        
        colIndexes = [i for x in zipCols for i, x2 in enumerate(colTake) if x == x2] 
    
        for cur,nxt in zip(zipCols, zipCols[1:]):
           if(skipChar == False):
               colChars.append(cur + nxt)
               skipChar = True
           else:
               skipChar = False
    
        for cur, nxt in zip(colIndexes, colIndexes[1:]):
            if(skipInd == False): #skips first iteration
                zippedData = list_zipper(colData[cur], colData[nxt])
                disposables += [colData[cur], colData[nxt]]
                colData.append(zippedData)
                skipInd = True
            else:
                skipInd = False
            
        
        #removes non-zipped version of data from main data
        for item in disposables:
            if item in colData:
                colData.remove(item)
    
        colChars = [char for char in colChars if char not in zipCols]

        print("Current column order and combined columns",colChars)

    
    
    def excel_writer(): #writes data to excel file
        wb = load_workbook('MasterFile.xlsx') #file to write data to
        ws = wb.active
        storeTo = input("Which column(s) to store data to: ").split()
    
        def col_writer(data, col): #writes to specific columns in an excel file and saves it
            max = ws.max_row
    
            for row, entry in enumerate(data, start=1):
                while(ws.cell(row = max, column=col).value is None):
                    if max == 1:
                        break
                    max -= 1
                ws.cell(row = row+max, column=col, value = entry)
    

        for data, col in zip(colData, storeTo):
                col = mt.col_to_num(col)
                col_writer(data, col)
                
        wb.save('MasterFile.xlsx')
        shutil.copy2("MasterFile.xlsx", "MasterFile_Backup.xlsx") #backup in case of failure

    col_extractor()
    if (len(zipCols) > 1 and len(zipCols) % 2 == 0):
        zip_cols(zipCols)
    elif (len(zipCols) % 2 != 0):
        print("Must have pairs of two")
    excel_writer()

excel_data_mover()